from selenium import webdriver
from selenium.webdriver.common.by import By 
from selenium.webdriver.common.keys import Keys
from openpyxl import load_workbook
from tkinter import Tk, Frame, Label, Button, messagebox

class aplicacao:
    # criando funções
    def __init__(self):

        # criando a aplicação
        self.layout = Tk()

        # titulo da aplicação
        self.layout.title("Previsão do tempo de São Paulo")

        # tamanho da tela
        self.layout.geometry("340x80")

        # layout da tela
        self.tela = Frame(self.layout)

        # mensagem na tela e botao
        self.descricao = Label(self.tela, text="Atualizar previsão na planilha:")
        self.exportar = Button(self.tela, text="Buscar previsão", command=self.buscar_previsao)

        # posicionamento da tela
        self.tela.pack()
        self.descricao.pack()
        self.exportar.pack()

        self.layout.mainloop()

    #pesquisando a previsao no navegador
    def buscar_previsao(self):

        link = webdriver.Chrome()

        link.get('https://www.google.com')

        link.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys('previsão do tempo em são paulo')
        link.find_element(By.XPATH, '/html/body/div[1]/div[3]/form/div[1]/div[1]/div[4]/center/input[1]').send_keys(Keys.ENTER)

        data_hora = link.find_element(By.XPATH, '//*[@id="wob_dts"]').text
        print(f'Data atual e hora atual: {data_hora}')

        tempo = link.find_element(By.XPATH, '//*[@id="wob_tm"]').text
        print(f'Temperatura atual: {tempo}')

        umidade = link.find_element(By.XPATH, '//*[@id="wob_wc"]/div[1]/div[2]/div[2]').text
        print(f'Umidade do ar atual: {umidade}')

        link.quit()

        # acessando o arquivo e salvando dados
        excel = load_workbook('Historico_Tempo.xlsx')
        plan = excel['tempo']

        # Salvando os dados
        next_row = 2
        for row in range(2, plan.max_row + 1):
            if plan.cell(row=row, column=1).value is None:
                next_row = row
                break
        else:
            next_row = plan.max_row + 1  
        
        plan.cell(row=next_row, column=1).value = data_hora
        plan.cell(row=next_row, column=2).value = tempo
        plan.cell(row=next_row, column=3).value = umidade

        excel.save('Historico_Tempo.xlsx')
        messagebox.showinfo("Atualização", "Dados salvos com sucesso no arquivo 'Historico_Tempo.xlsx'.")
          

tl = aplicacao()
                